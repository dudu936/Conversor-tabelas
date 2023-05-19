from pandas.core.window.expanding import Axis
import tabula as pdf
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_table(table, file, sheet_name="Inserção"):
    try:
        wb = load_workbook(file)
    except Exception as error:
        print(f"Não foi possível carregar o arquivo {file} devido ao erro {type(error).__name__}")
        wb = Workbook()
    ws = wb.active
    ws = wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    for row in dataframe_to_rows(table, index=False, header=True):
        ws.append(row)
    wb.save(file)

def remove_rows(table, rows_to_remove=[],index_to_name=0,):
    table.columns = table.iloc[index_to_name]
    for index, row in table.iterrows():
        values = row.values
        for value in values:
            for remove_value in rows_to_remove:
                if remove_value in str(value):
                    try:
                        table = table.drop(index)
                    except:
                      pass
    table = table.dropna(axis=1, how='all')
    return table

    
def separator_columns(tables, separator=[], index_to_remove=0):
    for column in separator:
        try:
            split_columns = tables[column].apply(lambda x: pd.Series(x.split(" ")))
            tables = pd.concat([tables, split_columns], axis=1).drop(columns=column)
        except:
            pass
    try:
      tables.columns = tables.iloc[index_to_remove - index_to_remove]
      tables = tables.drop(index_to_remove)
    except:
      pass
    return tables

def remove_columns(table, columns_to_remove=[]):
    for column in columns_to_remove:
        try:
            table = table.drop(column, axis=1)
        except:
            pass
    return table

def main(pdf_file="relatorio.pdf", xlsx_file="relatorio.xlsx", index=0, print_display=False, ignore=False,
         columns=[], rows=[],separator=[]):
    tables = pdf.read_pdf(pdf_file, pages="all")
    try:
      if not columns and not rows and not separator and not index:
        for table in tables:
          if print_display:
            display(table)
          if ignore:
            create_excel_table(table, xlsx_file)
          else:
             raise
      else:
          for table in tables:
            if rows:
              table = remove_rows(table, rows_to_remove=rows,index_to_name=index)
            if separator_columns:
              table = separator_columns(table, separator=separator, index_to_remove=index)
            if columns:
              table = remove_columns(table, columns_to_remove=columns)
            create_excel_table(table, xlsx_file)
            if print_display:
                display(table)
          print(f"Conversão do arquivo {pdf_file} para Excel concluída.")
    except Exception as error:
      print(f"Não foi possivel fazer a conversão! Ocorreu um erro do tipo {type(error).__name__}")
